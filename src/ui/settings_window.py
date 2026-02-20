#!./venv/bin/python3
"""
Settings window for Zedulo Payslips (Tkinter version).
Mousewheel works over all areas - scrollbar, labels, entries, buttons, empty space.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from src.config_manager import ConfigManager
from openpyxl import load_workbook


class SettingsWindow(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Settings")
        self.geometry("1000x750")
        self.resizable(True, False)

        self.attributes('-topmost', True)
        self.focus_force()

        self.config_manager = ConfigManager()
        self.config = self.config_manager.load()
        self.entries = {}

        self._create_scrollable_area()
        self._build_ui()

        # ✅ Bind mousewheel globally when window opens
        self._bind_mousewheel()

        self.after(200, lambda: self.attributes('-topmost', False))
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _create_scrollable_area(self):
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True)

        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollbar.pack(side="right", fill="y")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.scrollable_frame = tk.Frame(self.canvas)

        # Update scrollregion when frame resizes
        self.scrollable_frame.bind("<Configure>", self._on_frame_configure)
        # Update content width when canvas resizes
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

    def _bind_mousewheel(self):
        """Bind mousewheel globally for this window."""
        self.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        self.bind_all("<Button-4>", self._on_mousewheel, add="+")
        self.bind_all("<Button-5>", self._on_mousewheel, add="+")

    def _unbind_mousewheel(self):
        """Unbind mousewheel globally."""
        self.unbind_all("<MouseWheel>")
        self.unbind_all("<Button-4>")
        self.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        """
        Handle mousewheel by checking if scrollbar Y value actually changes.
        Prevents jitter at top/bottom edges.
        """
        # 1. Get scrollbar position BEFORE scroll attempt
        before_first, before_last = self.scrollbar.get()

        # 2. Attempt the scroll
        if event.num == 5:
            if before_last == 1.0:
                return

            self.canvas.yview_scroll(1, "units")  # DOWN
        elif event.num == 4:
            self.canvas.yview_scroll(-1, "units")  # UP

    def _on_frame_configure(self, event=None):
        """Update scrollregion when content changes size."""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        """Resize content to match canvas width."""
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_close(self):
        """Clean up bindings and close window."""
        self._unbind_mousewheel()
        self.destroy()

    def _build_ui(self):
        pad = 5
        parent = self.scrollable_frame

        for row, key in enumerate(self.config.keys()):
            label_text = key.replace("_", " ").title()

            lbl = tk.Label(parent, text=label_text)
            lbl.grid(row=row, column=0, sticky="w", padx=pad, pady=(pad, 0))
            # ✅ Bind mousewheel to label so it passes through
            lbl.bind("<MouseWheel>", self._on_mousewheel, add="+")
            lbl.bind("<Button-4>", self._on_mousewheel, add="+")
            lbl.bind("<Button-5>", self._on_mousewheel, add="+")

            entry = tk.Entry(parent, width=50)
            entry.grid(row=row, column=1, sticky="ew", padx=pad, pady=(pad, 0))
            entry.insert(0, self.config[key])
            self.entries[key] = entry
            # ✅ Bind mousewheel to entry so it passes through
            entry.bind("<MouseWheel>", self._on_mousewheel, add="+")
            entry.bind("<Button-4>", self._on_mousewheel, add="+")
            entry.bind("<Button-5>", self._on_mousewheel, add="+")

            if "FILEPATH" in key:
                btn = tk.Button(parent, text="Browse...", command=lambda k=key: self._browse_file(k))
                btn.grid(row=row, column=2, padx=pad, pady=(pad, 0))
                # ✅ Bind mousewheel to button so it passes through
                btn.bind("<MouseWheel>", self._on_mousewheel, add="+")
                btn.bind("<Button-4>", self._on_mousewheel, add="+")
                btn.bind("<Button-5>", self._on_mousewheel, add="+")
            elif "FOLDER" in key:
                btn = tk.Button(parent, text="Browse...", command=lambda k=key: self._browse_dir(k))
                btn.grid(row=row, column=2, padx=pad, pady=(pad, 0))
                # ✅ Bind mousewheel to button so it passes through
                btn.bind("<MouseWheel>", self._on_mousewheel, add="+")
                btn.bind("<Button-4>", self._on_mousewheel, add="+")
                btn.bind("<Button-5>", self._on_mousewheel, add="+")

        parent.columnconfigure(0, weight=0)
        parent.columnconfigure(1, weight=1)
        parent.columnconfigure(2, weight=0)

        save_btn = tk.Button(parent, text="Save", width=15, command=self._save)
        save_btn.grid(row=row + 1, column=1, sticky="e", pady=(20, 10), padx=pad)

        cancel_btn = tk.Button(parent, text="Cancel", width=15, command=self.destroy)
        cancel_btn.grid(row=row + 1, column=2, sticky="w", pady=(20, 10), padx=pad)

        parent.grid_rowconfigure(row + 1, weight=1)

        # Force scrollregion update after UI is built
        self.after_idle(self._on_frame_configure)

    def _browse_file(self, key):
        path = filedialog.askopenfilename(
            title="Select file",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
            parent=self
        )
        if path:
            self.entries[key].delete(0, tk.END)
            self.entries[key].insert(0, path)

    def _browse_dir(self, key):
        path = filedialog.askdirectory(title="Select Directory", parent=self)
        if path:
            self.entries[key].delete(0, tk.END)
            self.entries[key].insert(0, path)

    def _save(self):
        new_config = {}
        for key, entry in self.entries.items():
            val = entry.get().strip()

            if "FILEPATH" in key and val and not Path(val).exists():
                messagebox.showerror("Invalid Path", f"{val} does not exist!", parent=self)
                return
            elif "DIR" in key and val and not Path(val).is_dir():
                messagebox.showerror("Invalid Directory", f"{val} is not a directory!", parent=self)
                return
            elif "HEADER" in key.upper():
                if not val:
                    continue
                employee_sheet_path = None
                if "EMPLOYEE_SPREADSHEET_FILEPATH" in self.entries:
                    employee_sheet_path = self.entries["EMPLOYEE_SPREADSHEET_FILEPATH"].get()
                else:
                    employee_sheet_path = self.config["EMPLOYEE_SPREADSHEET_FILEPATH"]

                if not employee_sheet_path or not Path(employee_sheet_path).exists():
                    messagebox.showerror("Invalid Source", "Employee file path is not set or does not exist!", parent=self)
                    return

                try:
                    wb = load_workbook(employee_sheet_path)
                    header_found_in_all_sheets = True
                    missing_sheets = []

                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        row1_values = [str(cell.value).strip() for cell in sheet[1] if cell.value is not None]
                        if val not in row1_values:
                            header_found_in_all_sheets = False
                            missing_sheets.append(sheet_name)

                    if not header_found_in_all_sheets:
                        messagebox.showerror(
                            "Invalid Header",
                            f"\"{val}\" not found in row 1 of every sheet!\n\nMissing from: {', '.join(missing_sheets)}",
                            parent=self
                        )
                        return
                    wb.close()
                except Exception as e:
                    messagebox.showerror("Error Validating Header", str(e), parent=self)
                    return
            elif "CELL" in key:
                if not val:
                    continue
                try:
                    template_path = self.entries["PAYSLIP_TEMPLATE_FILEPATH"].get()
                    if template_path and Path(template_path).exists():
                        load_workbook(template_path).active[val]
                    else:
                        load_workbook(self.config["PAYSLIP_TEMPLATE_FILEPATH"]).active[val]
                except Exception as e:
                    messagebox.showerror("Invalid Cell reference", f"\"{val}\" location in the template sheet could not be validated!\nError:\n{e}", parent=self)
                    return

            new_config[key] = val

        try:
            self.config_manager.save(new_config)
            self.destroy()
            messagebox.showinfo("Settings Saved", "Configuration updated successfully.", parent=self.master)
        except Exception as e:
            messagebox.showerror("Error Saving", str(e), parent=self)
