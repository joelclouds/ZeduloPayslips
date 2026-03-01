import tkinter as tk
from tkinter import messagebox, filedialog
from threading import Thread, Event
from pathlib import Path
from src.config_manager import ConfigManager
from src.ui.settings_window import SettingsWindow
from src.services.file_explorer import open_with_default_app
from src.services.payslip_generator import PayslipGenerator
from openpyxl import load_workbook
from datetime import datetime
from src.services.mailing import send_payslip_email, send_bulk_payslips

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Zedulo Payslips Utility")
        self.root.geometry("1000x900")

        self.config_manager = ConfigManager()
        self.config = self.config_manager.load()

        self.generated_payslips = {}
        self.payslip_widgets = {}

        # Month selection: Dict of StringVars keyed by month name
        self.month_vars = {}
        self._build_ui()
        self.settings_window = None
        self.batch_mailing_inprog = False
        self.mailing_stop_event = Event()

    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        main_frame = tk.Frame(self.root, padx=40, pady=40)
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.columnconfigure(0, weight=1)

        # Title
        tk.Label(
            main_frame,
            text=f"Welcome {self.config['USERNAME']},",
            font=("Helvetica", 32, "bold")
        ).grid(row=0, column=0, pady=(0, 20))

        # Subtitle
        tk.Label(
            main_frame,
            text="Shall we start the payslip generation process now?\n...\nSimply provide the month(s), then data and the template",
            font=("Helvetica", 14)
        ).grid(row=1, column=0, pady=(0, 30))

        # --- Month Checklist Section ---
        month_frame = tk.LabelFrame(main_frame, text="Select Month(s)", padx=20, pady=10)
        month_frame.grid(row=2, column=0, sticky="ew", pady=(0, 30))
        month_frame.columnconfigure(tuple(range(6)), weight=1)

        # Month mapping: (Display Name, Backend Number)
        self.months_mapping = [
            ("January", 1), ("February", 2), ("March", 3), ("April", 4), ("May", 5), ("June", 6),
            ("July", 7), ("August", 8), ("September", 9), ("October", 10), ("November", 11), ("December", 12)
        ]

        self.month_checkboxes = {}
        self.month_vars = {}  # Keyed by month NUMBER (for backend)
        for idx, (month_name, month_num) in enumerate(self.months_mapping):
            row = idx // 6
            col = idx % 6

            var = tk.BooleanVar(value=False)
            self.month_vars[month_num] = var  # Key by number for backend

            cb = tk.Checkbutton(
                month_frame,
                text=month_name,  # UI shows NAME
                variable=var,
                font=("Helvetica", 10),
                anchor="w"
            )
            cb.grid(row=row, column=col, sticky="w", padx=10, pady=5)
            self.month_checkboxes[month_num] = cb

        # Quick select buttons
        quick_frame = tk.Frame(month_frame)
        quick_frame.grid(row=2, column=0, columnspan=6, pady=(10, 0))

        tk.Button(
            quick_frame,
            text="Select All",
            width=10,
            command=self._select_all_months
        ).pack(side="left", padx=5)

        tk.Button(
            quick_frame,
            text="Clear All",
            width=10,
            command=self._clear_all_months
        ).pack(side="left", padx=5)

        # Set current month as default selected
        current_month_num = datetime.now().month
        self.month_vars[current_month_num].set(True)

        # Buttons
        button_frame = tk.Frame(main_frame)
        button_frame.grid(row=3, column=0)

        self.generate_btn = tk.Button(button_frame, text="Generate Payslips", width=20, command=self._start_generation)
        self.generate_btn.grid(row=0, column=0, padx=10)

        tk.Button(button_frame, text="Settings", width=20, command=self._open_settings).grid(row=0, column=1, padx=10)

        # ← ADD SEND ALL BUTTON
        self.send_all_btn = tk.Button(button_frame, text="Send All Emails", width=20, command=self._send_all_emails, state="disabled")
        self.send_all_btn.grid(row=0, column=2, padx=10)
        # Progress label
        self.progress_label = tk.Label(main_frame, text="_", font=("Helvetica", 12))
        self.progress_label.grid(row=4, column=0, pady=(0, 20))

        # Scrollable payslip list section - DYNAMIC RESIZING
        list_frame = tk.LabelFrame(self.root, text="Generated Payslips", padx=10, pady=10)
        list_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=(0, 40))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)  # ← Canvas expands vertically

        self.payslip_canvas = tk.Canvas(list_frame, highlightthickness=0)
        self.payslip_scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.payslip_canvas.yview)
        self.payslip_scrollable_frame = tk.Frame(self.payslip_canvas)

        # Update scrollregion when canvas OR frame resizes
        self.payslip_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.payslip_canvas.configure(scrollregion=self.payslip_canvas.bbox("all"))
        )
        self.payslip_canvas.bind(
            "<Configure>",
            lambda e: self.payslip_canvas.itemconfig(self.canvas_window, width=e.width)
        )

        self.canvas_window = self.payslip_canvas.create_window((0, 0), window=self.payslip_scrollable_frame, anchor="nw", tags="scrollable_frame")
        self.payslip_canvas.configure(yscrollcommand=self.payslip_scrollbar.set)

        self.payslip_canvas.pack(side="left", fill="both", expand=True)
        self.payslip_scrollbar.pack(side="right", fill="y")

        # Mousewheel bindings (unchanged)
        self.payslip_canvas.bind("<Enter>", self._bind_payslip_mousewheel)
        self.payslip_canvas.bind("<Leave>", self._unbind_payslip_mousewheel)

        self.payslip_row = 0

    def _bind_payslip_mousewheel(self, event):
        self.payslip_canvas.bind_all("<MouseWheel>", self._on_payslip_mousewheel)
        self.payslip_canvas.bind_all("<Button-4>", self._on_payslip_mousewheel)
        self.payslip_canvas.bind_all("<Button-5>", self._on_payslip_mousewheel)

    def _unbind_payslip_mousewheel(self, event):
        self.payslip_canvas.unbind_all("<MouseWheel>")
        self.payslip_canvas.unbind_all("<Button-4>")
        self.payslip_canvas.unbind_all("<Button-5>")

    def _on_payslip_mousewheel(self, event):
        if event.num == 5 or event.delta == -120:
            self.payslip_canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta == 120:
            self.payslip_canvas.yview_scroll(-1, "units")

    def _select_all_months(self):
        for var in self.month_vars.values():
            var.set(True)

    def _clear_all_months(self):
        for var in self.month_vars.values():
            var.set(False)

    def _get_selected_months(self):
        # Returns list of INTEGERS [1, 2, 3...] for backend
        return sorted([month_num for month_num, var in self.month_vars.items() if var.get()])

    def _get_month_name(self, month_num):
        # Helper to convert number back to name for UI display
        for name, num in self.months_mapping:
            if num == month_num:
                return name
        return ""

    def _open_settings(self):
        if self.settings_window and self.settings_window.winfo_exists():
            self.settings_window.destroy()

        self.settings_window = SettingsWindow(self.root)

    def _start_generation(self):
        # Validate Month Selection
        selected_months = self._get_selected_months()
        if not selected_months:
            messagebox.showerror("Error", "Please select at least one month!", parent=self.root)
            return

        self.config = self.config_manager.load()

        if not self.config.get("EMPLOYEE_SPREADSHEET_FILEPATH"):
            messagebox.showerror("Error", "Employee spreadsheet not configured!", parent=self.root)
            return
        try:
            load_workbook(self.config["EMPLOYEE_SPREADSHEET_FILEPATH"])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open spreadsheet:\n{str(e)}", parent=self.root)
            return

        if not self.config.get("PAYSLIP_TEMPLATE_FILEPATH"):
            messagebox.showerror("Error", "Payslip template not configured!", parent=self.root)
            return
        try:
            load_workbook(self.config["PAYSLIP_TEMPLATE_FILEPATH"])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open spreadsheet:\n{str(e)}", parent=self.root)
            return

        # Reset Dict and UI
        self.generated_payslips.clear()
        self.payslip_widgets.clear()
        for widget in self.payslip_scrollable_frame.winfo_children():
            widget.destroy()
        self.payslip_row = 0

        self.generate_btn.config(state="disabled")

        # Display month NAMES in UI progress
        month_names = [self._get_month_name(m) for m in selected_months]
        months_display = ", ".join(month_names) if len(month_names) <= 3 else f"{month_names[0]}... ({len(month_names)} months)"
        self.progress_label.config(text=f"Starting generation for {months_display}...")

        # Pass INTEGERS to backend
        Thread(target=self._generate_worker, args=(selected_months,), daemon=True).start()

    def _generate_worker(self, selected_months):
        try:
            def progress_callback(counter, total, name=None, email=None, month=None, payslip_filepath=None):
                if payslip_filepath:
                    self.generated_payslips[payslip_filepath] = {
                        "name": name or "Unknown",
                        "email": email or "",
                        "month": month or "",
                        "payslip_filepath": payslip_filepath
                    }
                    # Pass email to _add_payslip_entry
                    self.root.after(0, lambda p=payslip_filepath, n=name, m=month, e=email: self._add_payslip_entry(p, n, m, e))
                self.root.after(0, lambda: self.progress_label.config(text=f"Generated {counter} / {total} {month} payslips"))

            for month in selected_months:
                generator = PayslipGenerator(month_no=month, progress_callback=progress_callback, config=self.config)
                generator.generate_payslips()

            self.root.after(0, lambda: self._generation_complete(selected_months))
        except Exception as e:
            import logging
            logging.error(e, exc_info=True)
            self.root.after(0, lambda: messagebox.showerror("Error", str(e), parent=self.root))
            self.root.after(0, lambda: self.generate_btn.config(state="normal"))

    def _generation_complete(self, selected_months):
        count = len(self.generated_payslips)
        month_names = [self._get_month_name(m) for m in selected_months]
        months_display = ", ".join(month_names) if len(month_names) <= 3 else f"{len(month_names)} months"
        self.progress_label.config(text=f"All payslips generated for {months_display}. ({count} total)")
        self.generate_btn.config(state="normal")

        # ← Enable Send All button if there are emails
        has_emails = any(v.get("email") and v["email"].strip() for v in self.generated_payslips.values())
        self.send_all_btn.config(state="normal" if has_emails else "disabled")

        messagebox.showinfo("Success", f"All payslips have been generated successfully for {months_display}.\n\nTotal: {count}", parent=self.root)

    def _add_payslip_entry(self, path, name, month, email=None):
        if path in self.payslip_widgets:
            return

        frame = tk.Frame(self.payslip_scrollable_frame, pady=5, padx=5)
        frame.grid(row=self.payslip_row, column=0, sticky="ew", pady=2)
        frame.columnconfigure(0, weight=1)  # Label expands

        info_frame = tk.Frame(frame)
        info_frame.grid(row=0, column=0, sticky="w", padx=(0, 10))

        display_text = f"{name}"
        if month:
            display_text += f" ({month})"

        label = tk.Label(info_frame, text=display_text, font=("Helvetica", 10, "bold"), anchor="w", wraplength=400)
        label.pack(anchor="w")

        path_label = tk.Label(info_frame, text=Path(path).name, font=("Helvetica", 8), fg="gray", anchor="w", wraplength=400)
        path_label.pack(anchor="w")

        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=0, column=1, sticky="e")

        review_btn = tk.Button(btn_frame, text="Review PDF", width=10, command=lambda p=path: self._review_payslip(p))
        review_btn.pack(side="left", padx=2)

        open_folder_btn = tk.Button(btn_frame, text="Open Folder", width=10, command=lambda p=path: self._open_folder(p))
        open_folder_btn.pack(side="left", padx=2)

        email_btn = tk.Button(btn_frame, text="Send Email", width=10)
        email_btn.config(command=lambda p=path, e=email, n=name, m=month, b=email_btn: self._send_payslip_email(p, e, n, m, b))
        email_btn.pack(side="left", padx=2)

        if not email or not email.strip():
            email_btn.config(state="disabled", text="No Email")

        self.payslip_widgets[path] = frame
        self.payslip_row += 1

        self.root.after(10, lambda: self.payslip_canvas.configure(scrollregion=self.payslip_canvas.bbox("all")))

    def _review_payslip(self, path):
        if Path(path).exists():
            open_with_default_app(path)
        else:
            messagebox.showerror("Error", f"File not found: {path}", parent=self.root)

    def _open_folder(self, path):
        folder = str(Path(path).parent)
        if Path(folder).exists():
            open_with_default_app(folder)
        else:
            messagebox.showerror("Error", f"Folder not found: {folder}", parent=self.root)

    def _send_payslip_email(self, pdf_path, recipient_email, employee_name, month, btn_widget):
        """Non-blocking single email send. Disables button during operation."""

        def _worker():
            try:
                success, message = send_payslip_email(recipient_email, employee_name, month, pdf_path)
                # Schedule UI update on main thread
                self.root.after(0, lambda: _finish(success, message))
            except Exception as e:
                self.root.after(0, lambda: _finish(False, str(e)))

        def _finish(success, message):
            btn_widget.config(state="normal", text="Resend Email")  # Re-enable button
            if success:
                messagebox.showinfo("Email Ready", message, parent=self.root)
            else:
                messagebox.showerror("Email Failed", message, parent=self.root)

        btn_widget.config(state="disabled", text="Sending...")
        Thread(target=_worker, daemon=True).start()

    def _send_all_emails(self):
        """Toggle start/stop for batch emailing."""

        # --- Nested Helper: Worker ---
        def _worker(payslip_list, stop_event):
            success, failed, errors = 0, 0, []
            try:
                for i, item in enumerate(payslip_list):
                    if stop_event.is_set():
                        break  # ← Cancel requested

                    try:
                        ok, msg = send_payslip_email(
                            recipient_email=item["email"],
                            employee_name=item["name"],
                            month=item["month"],
                            pdf_path=item["pdf"]
                        )
                        if ok:
                            success += 1
                        else:
                            failed += 1
                            errors.append({"email": item["email"], "error": msg})
                    except Exception as e:
                        failed += 1
                        errors.append({"email": item["email"], "error": str(e)})

                    # Progress update
                    self.root.after(0, lambda c=i+1, t=len(payslip_list):
                        self.progress_label.config(text=f"{c} / {t} emails opened..."))
            finally:
                # ← ALWAYS runs (cancel, crash, or complete)
                self.root.after(0, lambda: _finalize(success, failed, errors))

        # --- Nested Helper: Finalizer ---
        def _finalize(success, failed, errors):
            self.batch_mailing_inprog = False
            self.send_all_btn.config(text="Send All Emails", state="normal")
            self.progress_label.config(text="_")

            msg = f"Opened: {success}\nFailed: {failed}"
            if errors:
                msg += "\n\nErrors:\n" + "\n".join([f"{e['email']}: {e['error']}" for e in errors[:5]])
            messagebox.showinfo("Email Summary", msg, parent=self.root)

        # --- Main Logic ---
        if self.batch_mailing_inprog:
            # Cancel Request
            self.mailing_stop_event.set()
            self.send_all_btn.config(text="Stopping...", state="disabled")
            return

        # Start Request
        payslip_list = [
            {"email": v["email"], "name": v["name"], "month": v["month"], "pdf": k}
            for k, v in self.generated_payslips.items()
            if v.get("email") and v["email"].strip()
        ]

        if not payslip_list:
            messagebox.showwarning("No Emails", "No payslips with valid email addresses found.", parent=self.root)
            return

        if not messagebox.askyesno(
            "Confirm Send",
            f"Open {len(payslip_list)} Thunderbird compose windows?",
            parent=self.root
        ):
            return

        self.batch_mailing_inprog = True
        self.mailing_stop_event.clear()
        self.send_all_btn.config(text="Cancel Mailing", state="normal")
        self.progress_label.config(text=f"0 / {len(payslip_list)} emails opened...")

        Thread(target=_worker, args=(payslip_list, self.mailing_stop_event), daemon=True).start()

# --------------------------
# Entry point
# --------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()

