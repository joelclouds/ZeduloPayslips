from src.config_manager import ConfigManager
from datetime import datetime
from src.services.tax_calc import ghana_tax_calculator
from src.services.db import YTD_Tracker
from openpyxl import load_workbook
from pathlib import Path
import shutil
import subprocess

settings = ConfigManager().load()

class Column_header:
    def __init__(self, **kwargs):
        self.employee_sheet = kwargs.get("spreadsheet", None)
        self.header         = kwargs.get("header", "").strip()
        self.column         = None
        self.column_index   = None
        self.search_header()

    def search_header(self):
        headers        = [cell.value.strip() for cell in self.employee_sheet[1]]
        header_columns = [cell.column for cell in self.employee_sheet[1]]

        if self.header in headers:
            chosen_index = headers.index(self.header)
            self.column  = header_columns[chosen_index]
            self.column_index = self.column -1
        else:
            self.column = None
            self.column_index = None
            return -1

def first_date_of_month(month_num):
    year = datetime.now().year
    return year, month_num, 1

def last_date_of_month(month_num):
    year = datetime.now().year
    assert month_num in range(1, 13)

    if month_num in [9, 4, 6, 11]:
        day = 30
    elif month_num == 2:
        leap_year = not year % 4
        day = 29 if leap_year else 28
    else:
        day = 31

    return year, month_num, day

class PayslipGenerator:
    def __init__(self, month_no, **kwargs):
        """
        Loads Employee spreadsheet.
        Sets general payslip info.
        Finds important Columm Headers based on config values.
        Iterates over every non-empty row of "employee data".
        Fills payslip template file.
        """
        self.progress_callback = kwargs.get("progress_callback", None)
        self.counter = 0
        self.total   = 0

        # Load Employee Spreadsheet
        self.employee_sheet_filepath = settings["EMPLOYEE_SPREADSHEET_FILEPATH"]
        self.month_no       = month_no
        self.month          = datetime(1970, month_no, 1).strftime("%B")
        self.employee_sheet = None
        self.employee_sheet_headers = None
        self.template_sheet_cells   = None
        self.start_datetime_str = datetime(*first_date_of_month(month_no)).strftime("%d/%m/%Y")
        self.end_datetime_str   = datetime(*last_date_of_month(month_no)).strftime("%d/%m/%Y")

        self.load_employee_sheet()
        self._init_employee_sheet_headers()
        self._init_template_sheet_cells()

    def _init_employee_sheet_headers(self):
        # Find (important) Column Headers
        self.employee_sheet_headers = {
            "name"          : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_NAME_HEADER"]),
            "staff_number"  : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_STAFF_NUMBER_HEADER"]),
            "email"         : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_EMAIL_HEADER"]),
            "tin"           : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_TIN_HEADER"]),
            "position"      : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_POSITION_HEADER"]),
            "department"    : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_DEPARTMENT_HEADER"]),
            "account_number": Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_ACCOUNT_NUMBER_HEADER"]),
            "gross_income"  : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_GROSS_INCOME_HEADER"]),
            "untaxed_bonus" : Column_header(spreadsheet=self.employee_sheet, header=settings["EMPLOYEE_UNTAXED_BONUS_HEADER"])
        }

    def _init_template_sheet_cells(self):
        self.template_sheet_cells = {
            "payslip_date": {
                "location": settings["TEMPLATE_PAYSLIP_DATE_CELL"],
                "value"   : f"Date: {datetime.now().strftime('%d/%m/%Y')}"
             },
            "payslip_period" : {
                "location": settings["TEMPLATE_PAYSLIP_PERIOD_CELL"],
                "value"   : f"{self.start_datetime_str} - {self.end_datetime_str}"
            },
            "payslip_number" : {
                "location": settings["TEMPLATE_PAYSLIP_NUMBER_CELL"],
                "value"   : f"ZED{self.month_no}"
            },
            "name" : {
                "location": settings["TEMPLATE_NAME_CELL"],
                "value"   : None
            },
            "staff_number" : {
                "location": settings["TEMPLATE_STAFF_NUMBER_CELL"],
                "value"   : None
            },
            "email" : {
                "location": settings["TEMPLATE_EMAIL_CELL"],
                "value"   : None
            },
            "tin" : {
                "location": settings["TEMPLATE_TIN_CELL"],
                "value"   : None
            },
            "position" : {
                "location": settings["TEMPLATE_POSITION_CELL"],
                "value"   : None
            },
            "department" : {
                "location": settings["TEMPLATE_DEPARTMENT_CELL"],
                "value"   : None
            },
            "account_number": {
                "location": settings["TEMPLATE_ACCOUNT_NUMBER_CELL"],
                "value"   : None
            },
            "gross_income"  : {
                "location": settings["TEMPLATE_GROSS_INCOME_CELL"],
                "value"   : None
            },
            "untaxed_bonus" : {
                "location": settings["TEMPLATE_UNTAXED_BONUS_CELL"],
                "value"   : None
            },

            "employee_ssf" : {
                "location": settings["TEMPLATE_EMPLOYEE_SSF_CELL"],
                "value"   : None
            },
            "income_tax"   : {
                "location": settings["TEMPLATE_INCOME_TAX_CELL"],
                "value"   : None
            },
            "tier_2"       : {
                "location": settings["TEMPLATE_TIER_2_CELL"],
                "value"   : None
            },
            "employer_ssf" : {
                "location": settings["TEMPLATE_EMPLOYER_SSF_CELL"],
                "value"   : None
            },
            "bonus_tax"    : {
                "location": settings["TEMPLATE_BONUS_TAX_CELL"],
                "value"   : None
            },
            "total_deductions" : {
                "location": settings["TEMPLATE_TOTAL_DEDUCTIONS_CELL"],
                "value"   : None
            },
            "total_contributions": {
                "location": settings["TEMPLATE_TOTAL_CONTRIBUTIONS_CELL"],
                "value"   : None
            },
            "total_income" : {
                "location": settings["TEMPLATE_TOTAL_INCOME_CELL"],
                "value"   : None
            },
            "ytd_tier_1" : {
                "location": settings["TEMPLATE_YTD_TIER_1_CELL"],
                "value"   : None
            },
            "ytd_tier_2" : {
                "location": settings["TEMPLATE_YTD_TIER_2_CELL"],
                "value"   : None
            },
            "ytd_gross_pay" : {
                "location": settings["TEMPLATE_YTD_GROSS_PAY_CELL"],
                "value"   : None
            },
            "net_income"   : {
                "location": settings["TEMPLATE_NET_INCOME_CELL"],
                "value"   : None
            }
        }

    def load_employee_sheet(self):
        try:
            wb = load_workbook(self.employee_sheet_filepath)
            self.employee_sheet = wb[self.month]
            return self.employee_sheet
        except Exception as e:
            return -1

    def employee_sheet_rows_iter(self):
        header = True
        for row in self.employee_sheet:
            if header:
                header = False
            else:
                yield row

    def generate_payslip(self, employee_sheet_row):
        employee_entry = self.employee_sheet_headers.copy()

        for k, v in employee_entry.items():
            employee_entry[k] = employee_sheet_row[v.column_index].value if v.column else None

        assert type(employee_entry["staff_number"]) == int, f"For {self.month}, {employee_entry['name']} has no proper {settings['EMPLOYEE_STAFF_NUMBER_HEADER']} (it must be a number)"
        assert employee_entry["gross_income"] is not None, f"For {self.month}, {employee_entry['name']} has no {settings['EMPLOYEE_GROSS_INCOME_HEADER']} in the employee spreadsheet at least put 0 there"
        assert employee_entry["untaxed_bonus"] is not None, f"For {self.month}, {employee_entry['name']} has no {settings['EMPLOYEE_UNTAXED_BONUS_HEADER']} in the employee spreadsheet at least put 0 there"

        employee_entry.update(ghana_tax_calculator(int(employee_entry["gross_income"] *100), int(employee_entry["untaxed_bonus"] *100)).items())

        ytd_tracker = YTD_Tracker()
        ytd_tracker.set_month_record(self.month_no, employee_entry)
        employee_entry.update(ytd_tracker.get_cumulative_ytd(self.month_no, employee_entry))

        # Convert monetary fields from pesewas to GHS for export
        for k in employee_entry.keys():
            if (k in list(ghana_tax_calculator(0,0).keys()) or
                "ytd" in k):
                employee_entry[k] /= 100

        payslip_details = self.template_sheet_cells.copy()

        for k,v in payslip_details.items():
            if "payslip" in k:
                continue

            if k == "account_number":
                v['value'] = f"ECOBANK: {employee_entry[k]}"
            else:
                v['value'] = employee_entry[k]

        payslip_xlsx_filepath = self.write_payslip_xlsx(payslip_details)
        assert payslip_xlsx_filepath, "Failed to create a payslip spreadshhet from the template"

        payslip_pdf_filepath = self.spreadsheet_to_pdf(payslip_xlsx_filepath)
        assert payslip_pdf_filepath, "Failed to convert payslip spreadsheet to pdf"

        return {
            "details"      : payslip_details,
            "xlsx_filepath": payslip_xlsx_filepath,
            "pdf_filepath" : payslip_pdf_filepath
        }

    def write_payslip_xlsx(self, payslip_details: dict):
        output_dir = Path(settings["EMPLOYEE_PAYSLIPS_FOLDER"]) / str(self.month)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{payslip_details['name']['value'].replace(' ', '_')}_{self.month}_Payslip.xlsx"

        shutil.copy2(settings["PAYSLIP_TEMPLATE_FILEPATH"], output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        for v in payslip_details.values():
            if v['location']:
                ws[v['location']] = v.get('value')

        wb.save(output_path)
        wb.close()

        return str(output_path)

    def spreadsheet_to_pdf(self, spreadsheet_filepath):
        spreadsheet_path = Path(spreadsheet_filepath)
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(spreadsheet_path.parent), str(spreadsheet_path)],
            capture_output=True
        )

        pdf_path = spreadsheet_path.with_suffix('.pdf')

        return str(pdf_path) if pdf_path.exists() else None

    def generate_payslips(self):
        employee_spreadsheet_rows = list( self.employee_sheet_rows_iter() )
        self.total += len(employee_spreadsheet_rows)

        for row in employee_spreadsheet_rows:
            payslip_info = self.generate_payslip(row)
            self.counter += 1

            if self.progress_callback:
                self.progress_callback(
                    counter=self.counter,
                    total=self.total,
                    name=payslip_info['details']['name']['value'],
                    email=payslip_info['details']['email']['value'],
                    month=self.month,
                    payslip_filepath=payslip_info['pdf_filepath']
                )
